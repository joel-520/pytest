# -*- coding: utf-8 -*-
# @Time : 2022/7/20 14:45
# @Author : joel
# @Email : 770546904@qq.com
# @File : test_turn.py
import requests
from openpyxl import load_workbook

# accessToken = ''
# def test_get_token():
#     url = "http://10.0.3.243:8088/api/getToken.htm"
#     params = {
#         "client_id": "rTd9DUIYKQ",
#         "client_secret": "NrL3EMcowTxmqQcYVepuEwd0ugNDzTRO",
#         "grant_type": "code",
#         "response_type": "json"
#     }
#
#     res = requests.get(url=url, params=params)
#
#     print(res.text)
#     accessToken = res.json()['accessToken']
#     accessToken = "Bearer" + " " + accessToken
#     print(accessToken)




def test_cases_result():
    #用例位置,读取用印
    file_path = r"../cases/【测试用例】3.6.5.xlsx"
    workbook = load_workbook(file_path)
    worksheet = workbook["用印重置接口测试用例"]
    #用例有多少行
    count = worksheet.max_row - 1
    print("测试用例的数量：",count)

    #把每一天用例信息打印出来  （循环遍历）
    for row in range(3,worksheet.max_row+1):
        print("读取excel数据：",row)

    #读取接口用例信息
        api_url = worksheet.cell(row,4).value
        api_method = worksheet.cell(row,5).value
        # api_headers = worksheet.cell(row, 6).value
        api_params = worksheet.cell(row,7).value
        expected_code = worksheet.cell(row,8).value
        expected_result = worksheet.cell(row,10).value

        print(api_url,"",api_method,"",api_params)

        #通过requests发送请求
        res = requests.request(
            url = api_url,
            method = api_method,
            params = api_params
        )
        print("响应状态码： ",res.status_code)
        print("响应内容： ",res.text)

        worksheet.cell(row,9).value = res.text

        #判断   【断言】   【状态码】 【响应内容】
        if res.status_code != expected_code or (expected_result is not None and expected_result != res.text):
            print("测试用例执行不通过！！！")
            worksheet.cell(row,10).value = "不通过"

        else:
            print("测试用例执行通过！！")
            worksheet.cell(row,10).value = "通过"


    result = workbook.save(r"../cases/【测试用例】3.6.5.xlsx")
# test_cases_result()
