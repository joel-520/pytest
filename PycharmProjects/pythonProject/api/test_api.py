
import requests
from openpyxl import load_workbook


def test_cases_result():
    #用例位置,读取用印
    file_path = r"../cases/获取token.xlsx"
    workbook = load_workbook(file_path)
    worksheet = workbook["Sheet"]
    #用例有多少行
    count = worksheet.max_row - 1
    print("测试用例的数量：",count)

    #把每一天用例信息打印出来  （循环遍历）
    for row in range(2,worksheet.max_row+1):
        print("读取excel数据：",row)

    #读取接口用例信息
        api_url = worksheet.cell(row,4).value
        api_method = worksheet.cell(row,5).value
        api_params = worksheet.cell(row,7).value
        expected_code = worksheet.cell(row,8).value
        expected_result = worksheet.cell(row,10).value

        print(api_url,"",api_method,"",api_params)

        #通过requests发送请求
        res = requests.request(
            url = api_url,
            method = api_method,
            params = api_params,
        )
        print("响应状态码： ",res.status_code)
        print("响应内容： ",res.text)

        worksheet.cell(row,9).value = res.text

        #判断   【断言】   【状态码】 【响应内容】
        if res.status_code != expected_code or (expected_result is not None and expected_result != res.text):
            print("测试用例执行不通过！！！")
            worksheet.cell(row,10).value = "不通过"

        else:
            print("测试用例执行通过！！")
            worksheet.cell(row,10).value = "通过"


    workbook.save(r"../cases/获取token.xlsx")
test_cases_result()


 




# def write_to_excel(path: str, sheetStr, info, data):
#
#     #     实例化一个workbook对象
#     workbook = openpyxl.Workbook()
#     # 激活一个sheet
#     sheet = workbook.active
#     # 为sheet设置一个title
#     sheet.title = sheetStr
#
#     # 添加表头（不需要表头可以不用加）
#     data.insert(0, list(info))
#     # 开始遍历数组
#     for row_index, row_item in enumerate(data):
#
#         for col_index, col_item in enumerate(row_item):
#             # 写入
#             sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)
#
#     # 写入excel文件 如果path路径的文件不存在那么就会自动创建
#     workbook.save(path)
#     print('写入成功')
#
# if __name__ == '__main__':
#     # 数据结构1 path 文件的路径
#     path = r'../results/result.xlsx'
#     # 数据结构1Excel 中sheet 的名字
#     sheetStr = '用例执行结果'
#
#     info = ['result']
#     # 数据结构1数据
#     writeData = ["pass"]
#
#     # 执行
#     write_to_excel(path, sheetStr, info, writeData)








