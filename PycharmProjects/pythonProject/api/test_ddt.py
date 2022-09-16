# -*- coding: utf-8 -*-
# @Time : 2022/7/20 15:41
# @Author : joel
# @Email : 770546904@qq.com
# @File : test_ddt.py

#简单使用
import openpyxl

if __name__ == '__main__':
    path = 'F:/case/test_case.xlsx'
    # 读取excel文件
    workbook = openpyxl.load_workbook(path)
    # 读取所有sheet
    sheet = workbook.get_sheet_names()
    # 获取某个sheet
    sheet = workbook[sheet[0]]
    # 获取某个cell的值
    cell_val = sheet.cell(row=2, column=2).value
    print(cell_val)


#用例管理
import openpyxl
class ExcelHandler:
    def __init__(self, file):
        self.file = file
    def open_excel(self, sheet_name):
        """打开Excel、获取sheet"""
        wb = openpyxl.load_workbook(self.file)
        # 获取sheet_name
        sheet = wb[sheet_name]
        return sheet
    def get_header(self, sheet_name):
        """获取header(表头)"""
        wb = self.open_excel(sheet_name)
        header = []
        # 遍历第一行
        for i in wb[1]:
            # 将遍历出来的表头字段加入列表
            header.append(i.value)
        return header
    def read_excel(self, sheet_name):
        """读取所有数据"""
        sheet = self.open_excel(sheet_name)
        rows = list(sheet.rows)
        data = []
        # 遍历从第二行开始的每一行数据
        for row in rows[1:]:
            row_data = []
            # 遍历每一行的每个单元格
            for cell in row:
                row_data.append(cell.value)
                # 通过zip函数将两个列表合并成字典
                data_dict = dict(zip(self.get_header(sheet_name),row_data))
            data.append(data_dict)
        return data
    @staticmethod
    def write_excel(file, sheet_name, row, cloumn,data):
        """Excel写入数据"""
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, cloumn).value = data
        wb.save(file)
        wb.close()
if __name__ == "__main__":
    # 以下为测试代码
    excel = ExcelHandler('../data/cases.xlsx')
    data = excel.read_excel('login')

#ddt使用
def ddt(cls):
    """
    Class decorator for subclasses of ``unittest.TestCase``.
    Apply this decorator to the test case class, and then
    decorate test methods with ``@data``.
    For each method decorated with ``@data``, this will effectively create as
    many methods as data items are passed as parameters to ``@data``.
    The names of the test methods follow the pattern
    ``original_test_name_{ordinal}_{data}``. ``ordinal`` is the position of the
    data argument, starting with 1.
    For data we use a string representation of the data value converted into a
    valid python identifier.  If ``data.__name__`` exists, we use that instead.
    For each method decorated with ``@file_data('test_data.json')``, the
    decorator will try to load the test_data.json file located relative
    to the python file containing the method that is decorated. It will,
    for each ``test_name`` key create as many methods in the list of values
    from the ``data`` key.
    """
    for name, func in list(cls.__dict__.items()):
        if hasattr(func, DATA_ATTR):
            for i, v in enumerate(getattr(func, DATA_ATTR)):
                test_name = mk_test_name(name, getattr(v, "__name__", v), i)
                test_data_docstring = _get_test_data_docstring(func, v)
                if hasattr(func, UNPACK_ATTR):
                    if isinstance(v, tuple) or isinstance(v, list):
                        add_test(
                            cls,
                            test_name,
                            test_data_docstring,
                            func,
                            *v
                        )
                    else:
                        # unpack dictionary
                        add_test(
                            cls,
                            test_name,
                            test_data_docstring,
                            func,
                            **v
                        )
                else:
                    add_test(cls, test_name, test_data_docstring, func, v)
            delattr(cls, name)
        elif hasattr(func, FILE_ATTR):
            file_attr = getattr(func, FILE_ATTR)
            process_file_data(cls, name, func, file_attr)
            delattr(cls, name)
    return cls
def unpack(func):
    """
    Method decorator to add unpack feature.
    """
    setattr(func, UNPACK_ATTR, True)
    return func
def data(*values):
    """
    Method decorator to add to your test methods.
    Should be added to methods of instances of ``unittest.TestCase``.
    """
    global index_len
    index_len = len(str(len(values)))
    return idata(values)


import unittest
import ddt


# 装饰类
@ddt.ddt
class DdtDemo(unittest.TestCase):
    def setUp(self):
        pass

    def tearDown(self):
        pass

    # 装饰方法
    @ddt.data(("15312344578", "12345678"), ("15387654321", "12345678"))
    @ddt.unpack
    def test_ddt(self, username, password):
        print(username, password)


if __name__ == '__main__':
    unittest.main(verbosity=2)



import unittest
from common.requests_handler import RequestsHandler
from common.excel_handler import ExcelHandler
import ddt
import json
@ddt.ddt
class TestLogin(unittest.TestCase):
    # 读取excel中的数据
    excel = ExcelHandler('../data/cases.xlsx')
    case_data = excel.read_excel('login')
    print(case_data)
    def setUp(self):
        # 请求类实例化
        self.req = RequestsHandler()
    def tearDown(self):
        # 关闭session管理器
        self.req.close_session()
    @ddt.data(*case_data)
    def test_login_success(self,items):
        # 请求接口
        res = self.req.visit(method=items['method'],url=items['url'],json=json.loads(items['payload']),
                             headers=json.loads(items['headers']))
        try:
            # 断言：预期结果与实际结果对比
            self.assertEqual(res['code'], items['expected_result'])
            result = 'Pass'
        except AssertionError as e:
            result = 'Fail'
            raise e
        finally:
            # 将响应的状态码，写到excel的第9列，即写入返回的状态码
            TestLogin.excel.write_excel("../data/cases.xlsx", 'login', items['case_id'] + 1, 9, res['code'])
            # 如果断言成功，则在第10行(测试结果)写入Pass,否则，写入Fail
            TestLogin.excel.write_excel("../data/cases.xlsx", 'login', items['case_id'] + 1, 10, result)
if __name__ == '__main__':
    unittest.main()